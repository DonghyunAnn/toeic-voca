#!/usr/bin/env python3
"""엑셀 교재에서 단어의 티어(필수 / 토익 만점 완성)를 읽어낸다.

시트는 DAY마다 블록 두 개로 나뉜다. 왼쪽이 '필수 어휘', 오른쪽이
'토익 만점 완성'이다. 블로그는 왼쪽만 예문과 함께 게시했다.
열 위치가 시트마다 조금씩 달라서 열을 고정하지 않고 훑는다.

    python3 xlsx_tier.py "~/Downloads/ETS 토익 기출 보카 Day1~30 _protected.xlsx"
"""

import argparse
import json
import re
from pathlib import Path

import openpyxl

LATIN = re.compile(r"[A-Za-z]")


def norm(word):
    s = word.lower().replace("-", " ").replace("/", " ").replace("'", " ").replace("’", " ")
    return re.sub(r"\s+", " ", re.sub(r"[^a-z ]", " ", s)).strip()


def read(path):
    """{(day, 정규화한 단어): 'core' | 'bonus'}"""
    wb = openpyxl.load_workbook(Path(path).expanduser(), data_only=True, read_only=True)
    tiers, counts = {}, {"core": 0, "bonus": 0}
    try:
        for day in range(1, 31):
            sheet = f"Day {day}"
            if sheet not in wb.sheetnames:
                continue
            blocks = {}
            for row in wb[sheet].iter_rows(values_only=True):
                cells = ["" if c is None else c for c in row]
                for i in range(len(cells) - 1):
                    n = cells[i]
                    if not (isinstance(n, (int, float))
                            or (isinstance(n, str) and n.strip().isdigit())):
                        continue
                    for off in (1, 2):                 # 번호 바로 뒤 또는 한 칸 건너
                        if i + off >= len(cells):
                            break
                        w = cells[i + off]
                        if isinstance(w, str) and w.strip() and LATIN.search(w):
                            blocks.setdefault(i, []).append(w.strip())
                            break
            for rank, col in enumerate(sorted(blocks)):
                tier = "core" if rank == 0 else "bonus"
                for w in blocks[col]:
                    tiers[(day, norm(w))] = tier
                    counts[tier] += 1
    finally:
        wb.close()
    return tiers, counts


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx")
    ap.add_argument("--out", default=str(Path(__file__).parent / "data" / "tiers.json"))
    args = ap.parse_args()

    tiers, counts = read(args.xlsx)
    payload = {f"{d}|{w}": t for (d, w), t in tiers.items()}
    Path(args.out).write_text(json.dumps(payload, ensure_ascii=False, indent=0), encoding="utf-8")
    print(f"티어 {len(tiers)}건 -> {args.out}")
    print(f"  필수 어휘 {counts['core']} / 토익 만점 완성 {counts['bonus']}")


if __name__ == "__main__":
    main()
